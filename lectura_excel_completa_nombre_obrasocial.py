import time
import requests
import pandas as pd
import json
from openpyxl import load_workbook

# Define constants
BASE_URL = "http://190.52.34.65/Notificaciones/Informacion/ConsultaPuco"
CHUNK_SIZE = 100
NUEVO_CICLO = 10000

def send_request(dni):
    try:
        response = requests.post(BASE_URL, data={"dni": dni})
        if response.status_code == 200:
            json_data = response.json()
            print(json_data)
            if json_data.get("Estado") == "OK" and "Items" in json_data:
                items = json_data["Items"]
                extracted_data = []
                for item in items:
                    if all(key in item for key in ["NroDocumento", "Nombre", "ObraSocial", "Siglas"]):
                        nro_documento = item["NroDocumento"]
                        nombre = item["Nombre"]
                        obra_social = item["ObraSocial"]
                        siglas = item["Siglas"]
                        extracted_data.append((nro_documento, nombre, obra_social, siglas))
                return extracted_data
            elif json_data.get("Estado") in ["NotFound", "Error"]:
                nro_documento = dni
                nombre = "--"
                obra_social = "--"
                siglas = "--"
                extracted_data = [(nro_documento, nombre, obra_social, siglas)]
                return extracted_data
    except Exception as e:
        print(f'Exception in send_request: {e}')
    return None

def main():
    # Load Excel file
    wb = load_workbook(filename='SAMICDNI.xlsx')
    ws = wb['Hoja1']

    cont = 0
    data = []
    processed_nros_documentos = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        try:
            dni = str(row[0])  # Assuming NRO_DOCUMENTO_PACIENTE is in the first column (column A)
            result = send_request(dni)
            if result:
                for item in result:
                    if item[0] not in processed_nros_documentos:  # Check if NroDocumento is not processed before
                        print("Processed:", item)
                        processed_nros_documentos.add(item[0])  # Add NroDocumento to processed set
                        data.append(item)
                cont += 1
        except Exception as e:
            print(f'Exception in main loop: {e}')
        finally:
            if cont % CHUNK_SIZE == 0 and data:
                # Update Excel file with extracted data
                for i, (nro_documento, nombre, obra_social, siglas) in enumerate(data, start=2):
                    ws.cell(row=i, column=2, value=nombre)       # Column B: NOMBRE_PACIENTE
                    ws.cell(row=i, column=3, value=obra_social)  # Column C: OBRAS_SOCIAL
                    ws.cell(row=i, column=4, value=siglas)       # Column D: SIGLAS
                # Save the updated Excel file
                wb.save("SAMICDNI.xlsx")
                data = []

            if cont % NUEVO_CICLO == 0:
                print('NUEVO CICLO XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
                time.sleep(10)

    # Save any remaining data
    if data:
        for i, (nro_documento, nombre, obra_social, siglas) in enumerate(data, start=2):
            ws.cell(row=i, column=2, value=nombre)       # Column B: NOMBRE_PACIENTE
            ws.cell(row=i, column=3, value=obra_social)  # Column C: OBRAS_SOCIAL
            ws.cell(row=i, column=4, value=siglas)       # Column D: SIGLAS
        wb.save("SAMICDNI.xlsx")

if __name__ == "__main__":
    main()