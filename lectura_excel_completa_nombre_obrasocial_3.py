import time
import requests
import pandas as pd
import json
from openpyxl import load_workbook

# Define constants
BASE_URL = "http://190.52.34.65/Notificaciones/Informacion/ConsultaPuco"
CHUNK_SIZE = 1000
NUEVO_CICLO = 10000
EXCEL_NAME = "SAMICDNI_2.xlsx"

def send_request(dni):
    try:
        response = requests.post(BASE_URL, data={"dni": dni})
        if response.status_code == 200:
            json_data = response.json()
            print(json_data)
            if json_data.get("Estado") == "OK" and "Items" in json_data:
                items = json_data["Items"]
                extracted_data = []
                for item in items:
                    if all(key in item for key in ["NroDocumento", "Nombre", "ObraSocial", "Siglas"]):
                        nro_documento = item["NroDocumento"]
                        nombre = item["Nombre"]
                        obra_social = item["ObraSocial"]
                        siglas = item["Siglas"]
                        extracted_data.append((nro_documento, nombre, obra_social, siglas,dni))
                return extracted_data
            elif json_data.get("Estado") in ["NotFound", "Error"]:
                nro_documento = dni
                nombre = "--"
                obra_social = "--"
                siglas = "--"
                extracted_data = [(nro_documento, nombre, obra_social, siglas,dni)]
                return extracted_data
    except Exception as e:
        print(f'Exception in send_request: {e}')
    return None

def get_last_nombre_filled(ws):
    # Get the last filled row for the "NOMBRE" column (assuming it's column B)
    last_row = ws.max_row
    for row in range(last_row, 1, -1):
        nombre = ws.cell(row=row, column=2).value
        if nombre:
            return row
    return 1  # Return 1 if no nombre is found, meaning the first row

def main():
    # Load Excel file
    wb = load_workbook(filename=EXCEL_NAME)
    ws = wb['Hoja1']

    start_row = get_last_nombre_filled(ws) + 1  # Start from the next row after the last filled "NOMBRE"
    cont = 0
    processed_nros_documentos = set()

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        try:
            dni = str(row[0])  # Assuming NRO_DOCUMENTO_PACIENTE is in the first column (column A)
            result = send_request(dni)
            if result:
                for item in result:
                    if item[0] not in processed_nros_documentos:  # Check if NroDocumento is not processed before
                        print("Processed:", item)
                        processed_nros_documentos.add(item[0])  # Add NroDocumento to processed set
                        # Write data to Excel
                        ws.cell(row=start_row, column=2, value=item[1])  # Column B: NOMBRE_PACIENTE
                        ws.cell(row=start_row, column=3, value=item[2])  # Column C: OBRAS_SOCIAL
                        ws.cell(row=start_row, column=4, value=item[3])  # Column D: SIGLAS
                        ws.cell(row=start_row, column=5, value=dni) #numero DNI
                        start_row += 1  # Move to the next row
                        cont += 1

                        if cont % CHUNK_SIZE == 0:
                            # Save the updated Excel file after every CHUNK_SIZE records
                            wb.save(EXCEL_NAME)
                            print("--------------GUARDANDO EN EXCEL-------------------------")

                if cont % NUEVO_CICLO == 0:
                    print('NUEVO CICLO XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
                    time.sleep(10)

        except Exception as e:
            print(f'Exception in main loop: {e}')

    # Save the final updated Excel file
    wb.save(EXCEL_NAME)
    print("--------------GUARDANDO EN EXCEL-------------------------")

if __name__ == "__main__":
    main()