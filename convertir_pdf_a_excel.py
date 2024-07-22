from PyPDF2 import PdfReader
import openpyxl

def extract_text_from_pdf_and_split_to_excel(pdf_file_path, output_excel_name):
    # Load the Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'PDF Text'

    # Open the PDF file
    with open(pdf_file_path, 'rb') as file:
        reader = PdfReader(file)
        
        # Iterate through each page of the PDF
        for page_num, page in enumerate(reader.pages, start=1):
            # Extract text from the page
            page_text = page.extract_text()
            
            # Split text into lines
            page_lines = page_text.split('\n')
            
            # Iterate through each line
            for line in page_lines:
                # Check if the line starts with a number
                if not line.strip().startswith(('0', '1', '2', '3', '4', '5', '6', '7', '8', '9')):
                    continue  # Skip lines that do not start with a number
                
                # Split the line into price, code, and description
                split_line = line.strip().split(' ', 2)
                
                # If the line cannot be split, or if it doesn't have 3 parts, skip it
                if len(split_line) != 3:
                    continue
                
                # Extract code, description, and price
                price, code, description = split_line
                
                # Write the split information into separate columns in the output workbook
                row_num = len(sheet['A']) + 1  # Find the next empty row in column A
                sheet.cell(row=row_num, column=1, value=code)
                sheet.cell(row=row_num, column=2, value=description)
                sheet.cell(row=row_num, column=3, value=price.replace('.', ','))  # Replacing dot with comma
    
    # Save the output workbook
    workbook.save(output_excel_name)

# Example usage
pdf_file_path = 'articulosxr.pdf'
output_excel_name = 'processed_data_2.xlsx'
extract_text_from_pdf_and_split_to_excel(pdf_file_path, output_excel_name)