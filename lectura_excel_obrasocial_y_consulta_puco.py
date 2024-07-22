import time
import requests
import pandas as pd
import json
from openpyxl import load_workbook

# Define constants
BASE_URL = "http://190.52.34.65/Notificaciones/Informacion/ConsultaPuco"
CHUNK_SIZE = 1000
NUEVO_CICLO = 100000
EXCEL_NAME = "NUM_DNI_PRES24_OS_B.xlsx"


def send_request(dni):
    try:
        response = requests.post(BASE_URL, data={"dni": dni})
        if response.status_code == 200:
            json_data = response.json()
            print(json_data)
            if json_data.get("Estado") == "OK" and "Items" in json_data:
                items = json_data["Items"]
                extracted_data = []
                for item in items:
                    if all(key in item for key in ["NroDocumento", "Nombre", "ObraSocial", "Siglas"]):
                        nro_documento = item["NroDocumento"]
                        nombre = item["Nombre"]
                        obra_social = item["ObraSocial"]
                        siglas = item["Siglas"]
                        extracted_data.append((nro_documento, nombre, obra_social, siglas,dni))
                return extracted_data
            elif json_data.get("Estado") in ["NotFound", "Error"]:
                nro_documento = dni
                nombre = "--"
                obra_social = "--"
                siglas = "--"
                extracted_data = [(nro_documento, nombre, obra_social, siglas,dni)]
                return extracted_data
    except Exception as e:
        print(f'Exception in send_request: {e}')
    return None

def get_min_max_numero_dni(ws):
    min_row = 2  # Assuming title is in row 1 and data starts from row 2
    max_row = ws.max_row
    for row in range(2, ws.max_row + 1):
        numero_documento = ws.cell(row=row, column=1).value  # Assuming NUM_DOCUMENTO_PACIENTE is in the first column (column A)
        if not numero_documento:  # If the cell is empty
            max_row = row - 1  # Update max_row to the previous row, which is the last non-empty row
            break
    return min_row, max_row

def main():
    # Load Excel file
    wb = load_workbook(filename=EXCEL_NAME)
    ws = wb['Hoja1']

    min_row, max_row = get_min_max_numero_dni(ws)  # Get the min and max rows

    cont = 0
    processed_nros_documentos = set()

    for row in range(min_row, max_row + 1):  # Loop through the range of rows from min_row to max_row
        try:
            numero_documento = str(ws.cell(row=row, column=1).value)  # Assuming NUM_DOCUMENTO_PACIENTE is in the first column (column A)
            result = send_request(numero_documento)
            if result:
                for item in result:
                    if item[0] not in processed_nros_documentos:  # Check if NroDocumento is not processed before
                        print("Processed:", item)
                        processed_nros_documentos.add(item[0])  # Add NroDocumento to processed set
                        # Write data to Excel
                        ws.cell(row=row, column=5, value=item[1])  # Column E: NOMBRE_PACIENTE
                        ws.cell(row=row, column=6, value=item[2])  # Column F: OBRAS_SOCIAL
                        ws.cell(row=row, column=7, value=item[3])  # Column G: SIGLAS
                        ws.cell(row=row, column=8, value=numero_documento)  # Column H: numero DNI
                        cont += 1

                        if cont % CHUNK_SIZE == 0:
                            # Save the updated Excel file after every CHUNK_SIZE records
                            wb.save(EXCEL_NAME)
                            print("--------------GUARDANDO EN EXCEL-------------------------")

                if cont % NUEVO_CICLO == 0:
                    print('NUEVO CICLO XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
                    time.sleep(10)

        except Exception as e:
            print(f'Exception in main loop: {e}')

    # Save the final updated Excel file
    wb.save(EXCEL_NAME)
    print("--------------GUARDANDO EN EXCEL-------------------------")

if __name__ == "__main__":
    main()