
import openpyxl

def split_excel_lines(excel_file_path, output_excel_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Create a new workbook to store the split data
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    # Iterate through each row in the input workbook
    for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Extract information from the line
        line = row[0].strip()  # Assuming the line is in the first column
        split_line = line.split(' ')
        if len(split_line) >= 2:
            price = split_line[0].replace('.', ',')  # Replace dot with comma
            rest = ' '.join(split_line[1:])
            code, description = rest.split(' ', 1)

            # Write the split information into separate columns in the output workbook
            output_sheet.cell(row=row_num, column=1, value=price)
            output_sheet.cell(row=row_num, column=2, value=code)
            output_sheet.cell(row=row_num, column=3, value=description)
        else:
            # If the line cannot be split because it doesn't follow the expected format, skip it
            continue

    # Save the output workbook with the specified name
    output_workbook.save(output_excel_name)

# Example usage
excel_file_path = 'extracted_text.xlsx'
output_excel_name = 'processed_data.xlsx'
split_excel_lines(excel_file_path, output_excel_name)