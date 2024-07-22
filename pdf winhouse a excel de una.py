from PyPDF2 import PdfReader
import openpyxl

def extract_text_from_pdf_and_split_to_excel(pdf_file_path, output_excel_name):
    # Load the Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'PDF Text'

    # Open the PDF file
    with open(pdf_file_path, 'rb') as file:
        reader = PdfReader(file)
        
        # Iterate through each page of the PDF
        for page_num, page in enumerate(reader.pages, start=1):
            # Extract text from the page
            page_text = page.extract_text()
            
            # Split text into lines
            page_lines = page_text.split('\n')
            
            # Iterate through each line
            for line in page_lines:
                # Check if the line starts with a number
                if not line.strip().startswith(('0', '1', '2', '3', '4', '5', '6', '7', '8', '9')):
                    continue  # Skip lines that do not start with a number
                
                # Split the line into code, description, color and price
                split_line = line.strip().split(' ', 3)
                
                # If the line cannot be split, or if it doesn't have 4 parts, skip it
                if len(split_line) != 4:
                    continue
                
                # Extract code, description, and price
                code, description, color, precio = split_line
                
                # Write the split information into separate columns in the output workbook
                row_num = len(sheet['A']) + 1  # Find the next empty row in column A
                sheet.cell(row=row_num, column=1, value=code)
                sheet.cell(row=row_num, column=2, value=description)
                sheet.cell(row=row_num, column=3, value=color)
                sheet.cell(row=row_num, column=4, value=precio.replace('.', ','))  # Replacing dot with comma
    
    # Save the output workbook
    workbook.save(output_excel_name)

# Example usage
pdf_file_path = 'Lista de precios Winhouse 2024.pdf'
output_excel_name = 'processed_data_22.xlsx'
extract_text_from_pdf_and_split_to_excel(pdf_file_path, output_excel_name)
print('Proceso Finalizado')