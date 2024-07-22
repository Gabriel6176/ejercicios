import openpyxl

# Load the Excel file
file_path = 'articulosxr.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Create a new workbook to store all data
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = 'Combined Data'

# Iterate through each sheet in the input workbook
for sheet_name in workbook.sheetnames:
    current_sheet = workbook[sheet_name]
    
    # Iterate through each row in the specified range and copy it to the output sheet
    for row in current_sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=2, values_only=True):
        output_sheet.append(row)

# Save the output workbook
output_workbook.save('combined_data.xlsx')